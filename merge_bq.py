# -*- coding: utf-8 -*-
"""
merge_bq_pages.py - 报批表PDF添加BQ页面
功能：
1. 读取总表获取EL编号+材料名+BQ编号
2. 在BQ PDF中搜索每个编号所在的页码（正则匹配）
3. 将报批表PDF与对应BQ整页合并
4. 输出文件名格式：EL-XXX 材料名.pdf

用法：
  python merge_bq_pages.py --zongbiao <总表> --bq <BQ.pdf> --input <报批表PDF目录> --output <输出目录>

依赖：openpyxl, fitz（PyMuPDF）
"""
import sys, io, os, re, argparse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
import fitz  # PyMuPDF

parser = argparse.ArgumentParser(description='报批表PDF合并BQ页')
parser.add_argument('--zongbiao', required=True, help='总表Excel路径')
parser.add_argument('--bq', required=True, help='BQ标书PDF路径（必须有文字层）')
parser.add_argument('--input', required=True, help='报批表PDF输入目录')
parser.add_argument('--output', required=True, help='合并后PDF输出目录')
parser.add_argument('--src-sheet', type=int, default=0, help='总表Sheet索引(默认0)')
parser.add_argument('--src-start', type=int, default=7, help='总表数据起始行(默认7)')
args = parser.parse_args()

os.makedirs(args.output, exist_ok=True)

# ─── 1. 读取总表 ───
print('=== 读取总表 ===')
wb = openpyxl.load_workbook(args.zongbiao, data_only=True)
ws = wb.worksheets[args.src_sheet]

bq_list = []
for r in range(args.src_start, ws.max_row + 1):
    bq   = ws.cell(r, 1).value
    el   = ws.cell(r, 2).value
    name = ws.cell(r, 3).value
    if bq is None and el is None:
        break
    bq_list.append({
        'bq':   str(bq).strip()   if bq   else '',
        'el':   str(el).strip()   if el   else '',
        'name': str(name).strip() if name else ''
    })
print(f'共 {len(bq_list)} 项材料')

# ─── 2. 分析BQ PDF，建立页码索引 ───
print('\n=== 分析BQ PDF ===')
doc = fitz.open(args.bq)
print(f'BQ PDF共 {doc.page_count} 页')

page_index = {}
for page_num in range(doc.page_count):
    text = doc[page_num].get_text()
    # 匹配 BQ 编号格式：1.1, 2.1.3, 1.1-a 等
    matches = re.findall(r'\b(\d+\.\d+(?:[-.]\d+)?)\b', text)
    for m in matches:
        if m not in page_index:
            page_index[m] = page_num + 1  # 转为1基页码
print(f'建立索引：{len(page_index)} 个BQ编号 → 页码')

# ─── 3. 匹配BQ页码 ───
print('\n=== 匹配BQ页码 ===')
unmatched = []
for item in bq_list:
    bq = item['bq']
    if bq in page_index:
        item['bq_page'] = page_index[bq]
    else:
        # 尝试去掉后缀（1.1.1-a → 1.1.1 → 1.1）
        parts = bq.rsplit('.', 1)
        found = False
        for n in range(len(parts), 0, -1):
            base = '.'.join(parts[:n])
            if base in page_index:
                item['bq_page'] = page_index[base]
                found = True
                break
        if not found:
            item['bq_page'] = None
            unmatched.append(item)
            print(f'  未找到: {item["el"]}  BQ={bq}')

if not unmatched:
    print('  全部匹配成功 ✓')

# ─── 4. 合并PDF ───
print('\n=== 合并PDF ===')
ok_count = skip_count = 0
for item in bq_list:
    el      = item['el']
    bq_page = item.get('bq_page')

    # 查找报批表PDF（两种文件名格式：EL-001.pdf 或 "EL-001 材料名.pdf"）
    baopiao_file = None
    for f in os.listdir(args.input):
        if f.lower().endswith('.pdf') and f.lower().startswith(el.lower()):
            baopiao_file = os.path.join(args.input, f)
            break
    if not baopiao_file:
        print(f'  跳过（无报批表PDF）: {el}')
        skip_count += 1
        continue
    if bq_page is None:
        print(f'  跳过（无BQ页）: {el}')
        skip_count += 1
        continue

    # 合并
    output = fitz.open()
    baopiao_doc = fitz.open(baopiao_file)
    output.insert_pdf(baopiao_doc)
    baopiao_doc.close()
    output.insert_pdf(doc, from_page=bq_page - 1, to_page=bq_page - 1)

    safe_name = re.sub(r'[\\/:*?"<>|]', '-', item['name'])
    output_name = f'{el} {safe_name}.pdf'
    output_path = os.path.join(args.output, output_name)
    output.save(output_path)
    output.close()

    print(f'  ✓ {el}  {item["name"][:18]}  → BQ p{bq_page}')
    ok_count += 1

doc.close()

print(f'\n完成：{ok_count} 个PDF成功，{skip_count} 个跳过')
print(f'输出目录: {args.output}')

# ─── 5. 输出汇总表 ───
summary_path = os.path.join(args.output, 'merge_summary.txt')
with open(summary_path, 'w', encoding='utf-8') as f:
    f.write(f'BQ PDF: {args.bq}\n')
    f.write(f'总表:   {args.zongbiao}\n')
    f.write(f'成功:   {ok_count}  跳过: {skip_count}\n\n')
    for item in bq_list:
        page = item.get('bq_page', '?')
        status = '✓' if page else '✗'
        f.write(f'{status}  {item["el"]}  {item["name"]}  → BQ p{page}\n')
print(f'汇总: {summary_path}')
